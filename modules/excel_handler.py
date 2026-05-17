"""
excel_handler.py — Read and write Excel templates for bulk operations
Handles bulk import of ledgers, vouchers, stock items, and employees.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
from typing import Optional


TALLY_GROUPS = [
    "Sundry Debtors",
    "Sundry Creditors",
    "Bank Accounts",
    "Cash-in-Hand",
    "Sales Accounts",
    "Purchase Accounts",
    "Duties & Taxes",
    "Indirect Expenses",
    "Indirect Incomes",
    "Direct Expenses",
    "Direct Incomes",
    "Capital Account",
    "Loans (Liability)",
    "Fixed Assets",
    "Current Assets",
    "Current Liabilities",
    "Investments",
    "Suspense A/c",
    "Provisions",
    "Reserves & Surplus",
    "Secured Loans",
    "Unsecured Loans",
    "Stock-in-Hand",
    "Deposits (Asset)",
    "Loans & Advances (Asset)",
    "Branch / Divisions",
    "Salary Payable",
    "TDS Payable",
]

INDIAN_STATES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu",
    "Telangana", "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Andaman & Nicobar Islands", "Chandigarh", "Dadra & Nagar Haveli",
    "Daman & Diu", "Delhi", "Jammu & Kashmir", "Ladakh", "Lakshadweep", "Puducherry",
]


# ─── Template Creators ───────────────────────────────────────────────────────

def create_ledger_template() -> bytes:
    """Create an Excel template for bulk ledger creation with dropdowns."""
    wb = openpyxl.Workbook()

    # ── Hidden sheet for dropdown source data ──
    ref_ws = wb.create_sheet("_Groups")
    ref_ws.sheet_state = "hidden"
    for i, g in enumerate(TALLY_GROUPS, start=1):
        ref_ws.cell(row=i, column=1, value=g)
    for i, s in enumerate(INDIAN_STATES, start=1):
        ref_ws.cell(row=i, column=2, value=s)

    # ── Main ledger sheet ──
    ws = wb.active
    ws.title = "Ledgers"

    headers = [
        "Ledger Name*",
        "Under Group*",
        "GSTIN (optional)",
        "State (optional)",
        "Opening Balance",
        "Bank Account No",
        "Bank Name",
        "IFSC Code",
        "Mobile",
        "Email",
        "Credit Limit",
        "Credit Days",
    ]
    sample = [
        "Ramesh Traders",
        "Sundry Debtors",
        "27AAACR1234A1Z5",
        "Bihar",
        "0",
        "",
        "",
        "",
        "9876543210",
        "ramesh@example.com",
        "50000",
        "30",
    ]

    _style_header_row(ws, headers)
    ws.append(sample)
    _auto_fit_columns(ws)

    # ── Dropdown: Under Group (column B = col 2) ──
    group_dv = DataValidation(
        type="list",
        formula1=f"_Groups!$A$1:$A${len(TALLY_GROUPS)}",
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Group",
        error="Please select a valid Tally group from the list.",
    )
    ws.add_data_validation(group_dv)
    group_dv.add(f"B2:B1000")

    # ── Dropdown: State (column D = col 4) ──
    state_dv = DataValidation(
        type="list",
        formula1=f"_Groups!$B$1:$B${len(INDIAN_STATES)}",
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(state_dv)
    state_dv.add("D2:D1000")

    # ── Instructions sheet ──
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "Smart Tally Agent — Ledger Import Template"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A3"] = "Column Guide:"
    ws2["A3"].font = Font(bold=True)
    instructions = [
        ("Ledger Name*",    "Required. Exact name as you want in Tally."),
        ("Under Group*",    "Required. Select from dropdown — do NOT type manually."),
        ("GSTIN",           "Optional. 15-character GSTIN for registered parties."),
        ("State",           "Optional. Select from dropdown."),
        ("Opening Balance", "Optional. Enter 0 if no opening balance."),
        ("Bank Account No", "Only for Bank Accounts group."),
        ("Bank Name",       "Only for Bank Accounts group."),
        ("IFSC Code",       "Only for Bank Accounts group."),
        ("Mobile",          "Optional. 10-digit mobile number."),
        ("Email",           "Optional. Party email address."),
        ("Credit Limit",    "Optional. 0 = no limit."),
        ("Credit Days",     "Optional. 0 = no limit."),
    ]
    for i, (col, desc) in enumerate(instructions, start=4):
        ws2.cell(row=i, column=1, value=col).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=desc)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55

    return _save_to_bytes(wb)


def create_sales_template() -> bytes:
    """Create an Excel template for bulk sales voucher entry."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Vouchers"

    headers = ["Date (YYYYMMDD)*", "Party Name*", "Total Amount*", "GST %*", "Narration", "Interstate (Yes/No)"]
    sample  = ["20240101", "Ramesh Traders", "11800", "18", "Sales of goods", "No"]

    _style_header_row(ws, headers)
    ws.append(sample)

    # Dropdown for Interstate
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("F2:F1000")

    _auto_fit_columns(ws)
    return _save_to_bytes(wb)


def create_purchase_template() -> bytes:
    """Create an Excel template for bulk purchase voucher entry."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Vouchers"

    headers = ["Date (YYYYMMDD)*", "Vendor Name*", "Total Amount*", "GST %*", "Narration", "Interstate (Yes/No)"]
    sample  = ["20240101", "Suresh Suppliers", "5900", "18", "Purchase of material", "No"]

    _style_header_row(ws, headers)
    ws.append(sample)

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("F2:F1000")

    _auto_fit_columns(ws)
    return _save_to_bytes(wb)


def create_employee_template() -> bytes:
    """Create an Excel template for bulk employee/payroll data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ["Employee Name*", "Employee ID*", "Department*", "Basic Salary*"]
    sample  = ["Rahul Kumar", "EMP001", "Accounts", "30000"]

    _style_header_row(ws, headers)
    ws.append(sample)
    _auto_fit_columns(ws)
    return _save_to_bytes(wb)


def create_payment_template() -> bytes:
    """Create an Excel template for bulk payment/receipt entries."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = ["Date (YYYYMMDD)*", "Party Name*", "Amount*", "Bank/Cash Ledger*", "Narration"]
    sample  = ["20240101", "Ramesh Traders", "50000", "Bank Account", "Payment against invoice"]

    _style_header_row(ws, headers)
    ws.append(sample)
    _auto_fit_columns(ws)
    return _save_to_bytes(wb)


# ─── Excel Readers ───────────────────────────────────────────────────────────

def read_ledger_excel(file_bytes) -> dict:
    """Read ledger data from uploaded Excel file — supports all new columns."""
    try:
        df = pd.read_excel(file_bytes, engine="openpyxl", sheet_name="Ledgers")
        df.columns = df.columns.str.strip().str.replace("*", "", regex=False)
        df = df.dropna(subset=["Ledger Name", "Under Group"])
        df = df.fillna("")

        records = []
        for _, row in df.iterrows():
            name = str(row.get("Ledger Name", "")).strip()
            if not name:
                continue
            records.append({
                "name":            name,
                "parent":          str(row.get("Under Group", "")).strip(),
                "gstin":           str(row.get("GSTIN (optional)", "")).strip().upper(),
                "state":           str(row.get("State (optional)", "")).strip(),
                "opening_balance": _safe_float(row.get("Opening Balance", 0)),
                "bank_ac_no":      str(row.get("Bank Account No", "")).strip(),
                "bank_name":       str(row.get("Bank Name", "")).strip(),
                "ifsc":            str(row.get("IFSC Code", "")).strip().upper(),
                "mobile":          str(row.get("Mobile", "")).strip(),
                "email":           str(row.get("Email", "")).strip(),
                "credit_limit":    _safe_float(row.get("Credit Limit", 0)),
                "credit_days":     int(_safe_float(row.get("Credit Days", 0))),
            })
        return {"success": True, "data": records, "count": len(records)}
    except Exception as e:
        return {"success": False, "error": f"Failed to read Excel file: {str(e)}"}


def read_sales_excel(file_bytes) -> dict:
    """Read sales voucher data from uploaded Excel file."""
    try:
        df = pd.read_excel(file_bytes, engine="openpyxl")
        df.columns = df.columns.str.strip().str.replace("*", "", regex=False)
        df = df.dropna(subset=["Date (YYYYMMDD)", "Party Name", "Total Amount"])
        df = df.fillna("")

        records = []
        for _, row in df.iterrows():
            records.append({
                "date":        str(int(float(row["Date (YYYYMMDD)"]))),
                "party":       str(row["Party Name"]).strip(),
                "amount":      float(row["Total Amount"]),
                "gst_pct":     float(row.get("GST %", 18)),
                "narration":   str(row.get("Narration", "")).strip(),
                "is_interstate": str(row.get("Interstate (Yes/No)", "No")).strip().lower() == "yes",
            })
        return {"success": True, "data": records, "count": len(records)}
    except Exception as e:
        return {"success": False, "error": f"Failed to read Excel file: {str(e)}"}


def read_employee_excel(file_bytes) -> dict:
    """Read employee data from uploaded Excel file."""
    try:
        df = pd.read_excel(file_bytes, engine="openpyxl")
        df.columns = df.columns.str.strip().str.replace("*", "", regex=False)
        df = df.dropna(subset=["Employee Name", "Basic Salary"])
        df = df.fillna("")

        records = []
        for _, row in df.iterrows():
            records.append({
                "name":         str(row.get("Employee Name", "")).strip(),
                "emp_id":       str(row.get("Employee ID", "")).strip(),
                "dept":         str(row.get("Department", "")).strip(),
                "basic_salary": float(row.get("Basic Salary", 0)),
            })
        return {"success": True, "data": records, "count": len(records)}
    except Exception as e:
        return {"success": False, "error": f"Failed to read Excel file: {str(e)}"}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _safe_float(val) -> float:
    try:
        return float(val) if val != "" else 0.0
    except (ValueError, TypeError):
        return 0.0


def _style_header_row(ws, headers: list):
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 15)


def _save_to_bytes(wb) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
