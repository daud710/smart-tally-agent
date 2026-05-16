"""
gstr3b_generator.py — Generate GSTR-3B summary in Excel format
GSTR-3B is the monthly summary return showing tax liability and ITC.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_gstr3b_excel(
    gstin: str,
    period: str,
    sales_data: list,
    purchase_data: list
) -> dict:
    """
    Generate a GSTR-3B summary Excel workbook.
    
    Args:
        gstin:          Company GSTIN
        period:         "March 2024" or "03/2024"
        sales_data:     list of sales with cgst, sgst, igst, taxable
        purchase_data:  list of purchases with cgst, sgst, igst, taxable (ITC)
    
    Returns:
        {"success": True, "bytes": bytes, "filename": str, "summary": dict}
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GSTR-3B"

        # ── Styles ──────────────────────────────────────────────────────────
        header_fill  = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        section_fill = PatternFill(start_color="ebf8ff", end_color="ebf8ff", fill_type="solid")
        total_fill   = PatternFill(start_color="e6fffa", end_color="e6fffa", fill_type="solid")
        thin_border  = Border(
            left   = Side(style="thin"),
            right  = Side(style="thin"),
            top    = Side(style="thin"),
            bottom = Side(style="thin")
        )

        def cell(row, col, value, bold=False, fill=None, align="left", num_format=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, color="FFFFFF" if fill == header_fill else "000000",
                          name="Calibri", size=10)
            if fill:
                c.fill = fill
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = thin_border
            if num_format:
                c.number_format = num_format
            return c

        # ── Title ───────────────────────────────────────────────────────────
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = f"GSTR-3B SUMMARY — {period}"
        c.font  = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        c.fill  = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:F2")
        ws["A2"].value = f"GSTIN: {gstin}"
        ws["A2"].font  = Font(bold=True, size=11, name="Calibri")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # ── Section 3.1 — Tax on Outward Supplies ───────────────────────────
        row = 4
        ws.merge_cells(f"A{row}:F{row}")
        cell(row, 1, "3.1  Tax on Outward Supplies (Sales)", bold=True, fill=header_fill, align="center")
        ws.row_dimensions[row].height = 22

        row += 1
        headers = ["Nature of Supplies", "Total Taxable Value (₹)",
                   "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
        for col, h in enumerate(headers, 1):
            cell(row, col, h, bold=True, fill=section_fill, align="center")
        ws.row_dimensions[row].height = 18

        # Outward supply totals
        total_taxable_out = round(sum(s.get("taxable", 0) for s in sales_data), 2)
        total_cgst_out    = round(sum(s.get("cgst",    0) for s in sales_data), 2)
        total_sgst_out    = round(sum(s.get("sgst",    0) for s in sales_data), 2)
        total_igst_out    = round(sum(s.get("igst",    0) for s in sales_data), 2)

        row += 1
        ws.cell(row, 1).value = "(a) Taxable supplies (other than zero rated/nil rated)"
        for col, val in enumerate([total_taxable_out, total_igst_out,
                                    total_cgst_out, total_sgst_out, 0], 2):
            cell(row, col, val, num_format="#,##0.00")

        row += 1
        ws.cell(row, 1).value = "(b) Zero rated supply (Export)"
        for col in range(2, 7):
            cell(row, col, 0, num_format="#,##0.00")

        # ── Section 4 — ITC Available ────────────────────────────────────────
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        cell(row, 1, "4.   Eligible ITC (Input Tax Credit)", bold=True, fill=header_fill, align="center")
        ws.row_dimensions[row].height = 22

        row += 1
        for col, h in enumerate(headers, 1):
            cell(row, col, h, bold=True, fill=section_fill, align="center")

        total_taxable_in = round(sum(p.get("taxable", 0) for p in purchase_data), 2)
        total_cgst_in    = round(sum(p.get("cgst",    0) for p in purchase_data), 2)
        total_sgst_in    = round(sum(p.get("sgst",    0) for p in purchase_data), 2)
        total_igst_in    = round(sum(p.get("igst",    0) for p in purchase_data), 2)

        row += 1
        ws.cell(row, 1).value = "(A) ITC Available — Inputs"
        for col, val in enumerate([total_taxable_in, total_igst_in,
                                    total_cgst_in, total_sgst_in, 0], 2):
            cell(row, col, val, num_format="#,##0.00")

        # ── Section 5.1 — Net Tax Payable ────────────────────────────────────
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        cell(row, 1, "5.1  Net Tax Payable", bold=True, fill=header_fill, align="center")
        ws.row_dimensions[row].height = 22

        net_cgst = round(total_cgst_out - total_cgst_in, 2)
        net_sgst = round(total_sgst_out - total_sgst_in, 2)
        net_igst = round(total_igst_out - total_igst_in, 2)
        total_tax_payable = round(net_cgst + net_sgst + net_igst, 2)

        row += 1
        for col, h in enumerate(["Description", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Total (₹)", ""], 1):
            cell(row, col, h, bold=True, fill=section_fill, align="center")

        row += 1
        ws.cell(row, 1).value = "Output Tax Liability"
        for col, val in enumerate([total_igst_out, total_cgst_out, total_sgst_out,
                                    round(total_igst_out + total_cgst_out + total_sgst_out, 2), ""], 2):
            cell(row, col, val, num_format="#,##0.00")

        row += 1
        ws.cell(row, 1).value = "Less: ITC Claimed"
        for col, val in enumerate([total_igst_in, total_cgst_in, total_sgst_in,
                                    round(total_igst_in + total_cgst_in + total_sgst_in, 2), ""], 2):
            cell(row, col, val, num_format="#,##0.00")

        row += 1
        ws.cell(row, 1).value = "Net Tax Payable"
        ws.cell(row, 1).font  = Font(bold=True)
        ws.cell(row, 1).fill  = total_fill
        for col, val in enumerate([net_igst, net_cgst, net_sgst, total_tax_payable, ""], 2):
            c = cell(row, col, val, bold=True, fill=total_fill, num_format="#,##0.00")

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 45
        for col_letter in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 18

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"GSTR3B_{gstin}_{period.replace(' ', '_')}.xlsx"
        summary = {
            "total_sales":        total_taxable_out,
            "output_cgst":        total_cgst_out,
            "output_sgst":        total_sgst_out,
            "output_igst":        total_igst_out,
            "itc_cgst":           total_cgst_in,
            "itc_sgst":           total_sgst_in,
            "itc_igst":           total_igst_in,
            "net_cgst_payable":   max(net_cgst, 0),
            "net_sgst_payable":   max(net_sgst, 0),
            "net_igst_payable":   max(net_igst, 0),
            "total_tax_payable":  max(total_tax_payable, 0),
        }
        return {"success": True, "bytes": buf.getvalue(),
                "filename": filename, "summary": summary}

    except Exception as e:
        return {"success": False, "error": f"GSTR-3B generation failed: {str(e)}"}
